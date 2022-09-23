# !python3 
# class_excel.py - This module build a workbook for an exel and insert the data inside this workbook.


# Modules
import json
from pprint import pprint
from typing import final
import openpyxl

# Loggin module
import logging
logging.basicConfig(filename="myProgramLog.log",level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')

class Excel:
    
    def __init__(self):
        self.name = ""
    
        
    # Processing data & create and Excel file   
    def processing_data(self, data):
        try:
            logging.info('Creating excel...')
            # Create our excel
            wb = openpyxl.Workbook() # Create a blank workbook
            
            
            # Create our sheet.
            sheet = wb.active
            sheet.title = 'harry_potter_characters'
            logging.info(f"Sheet name is: {sheet.title}")
            
            
            # Insert headers in our worksheet.
                # Take the keys of our paramethers data.
            header_keys_excel = [] #Store the keys of our JSON 
            for k in data[0].keys():
                header_keys_excel.append(k.title())
            #print(header_keys_excel) 
                       
            
            # Insert the headers.
            for index, value in enumerate(header_keys_excel):
                sheet.cell(1, index + 1).value = value 
            
            # Insert data in rows.
            hp_data = [] 
            for v in data:
                name = v['name']
                gender = v['gender']
                date_of_birth = v['date_of_birth']
                hair_colour = v['hair_colour']
                species = v['species']
                house = v['house']
                wizard = v['wizard']
                hogwarts_student = v['hogwarts_student']
                ancestry = v['ancestry']
                wand = str(v['wand'])
                image = str(v['image'])
                hp_data.append([name, gender, date_of_birth, hair_colour, species, house, wizard, hogwarts_student, ancestry, wand, image])
            #pprint(hp_data)
            for i in range(len(data)):
                for index, value in enumerate(hp_data[i]):
                    sheet.cell( i + 2  , index + 1 , value = value)
            logging.info('Insert data correctly')
            
            
            # Save our workbook
            wb.save('excel/harry_potter_analysis.xlsx') #  Change the title and path
            logging.info('Woorkbook saved!')
            
        except:
            logging.error('Err creating the excel')
        finally:
            logging.info('Excel create succesfully!')
              
        
        
        
        
